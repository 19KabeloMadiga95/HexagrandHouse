from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


# =========================================================
# PROJECT PATHS
# =========================================================

BASE_DIR = Path(__file__).resolve().parents[3]

BACKTEST_DIR = (
    BASE_DIR
    / "data"
    / "exports"
    / "backtesting"
)

OUTPUT_FILE = (
    BACKTEST_DIR
    / "unified_model_performance_dashboard.xlsx"
)


# =========================================================
# INPUT FILES
# =========================================================

BACKTEST_FILES = [
    {
        "GameFamily": "PowerBall",
        "File": BACKTEST_DIR / "powerball_model_comparison_backtest.xlsx",
    },
    {
        "GameFamily": "Lotto",
        "File": BACKTEST_DIR / "lotto_model_comparison_backtest.xlsx",
    },
    {
        "GameFamily": "Daily Lotto",
        "File": BACKTEST_DIR / "daily_lotto_model_comparison_backtest.xlsx",
    },
    {
        "GameFamily": "UK49s",
        "File": BACKTEST_DIR / "uk49s_model_comparison_backtest.xlsx",
    },
]


# =========================================================
# HELPERS
# =========================================================

def read_excel_sheet(path, sheet_name):
    if not path.exists():
        return pd.DataFrame()

    try:
        return pd.read_excel(
            path,
            sheet_name=sheet_name,
            engine="openpyxl"
        )

    except Exception:
        return pd.DataFrame()


def load_all_summaries():
    frames = []

    missing_files = []

    for item in BACKTEST_FILES:
        game_family = item["GameFamily"]
        file_path = item["File"]

        if not file_path.exists():
            missing_files.append({
                "GameFamily": game_family,
                "ExpectedFile": str(file_path),
                "Status": "Missing",
            })
            continue

        summary = read_excel_sheet(
            file_path,
            "Summary"
        )

        if summary.empty:
            missing_files.append({
                "GameFamily": game_family,
                "ExpectedFile": str(file_path),
                "Status": "Summary sheet missing or empty",
            })
            continue

        summary.insert(
            0,
            "GameFamily",
            game_family
        )

        frames.append(summary)

    if frames:
        combined = pd.concat(
            frames,
            ignore_index=True
        )

    else:
        combined = pd.DataFrame()

    missing_df = pd.DataFrame(
        missing_files
    )

    return combined, missing_df


def safe_numeric(df, col):
    if col in df.columns:
        df[col] = pd.to_numeric(
            df[col],
            errors="coerce"
        )

    return df


# =========================================================
# DASHBOARD TABLES
# =========================================================

def build_dashboard_summary(combined_df, missing_df):
    rows = [
        {
            "Metric": "Report Generated At",
            "Value": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        },
        {
            "Metric": "Games Loaded",
            "Value": (
                combined_df["GameFamily"].nunique()
                if not combined_df.empty
                else 0
            ),
        },
        {
            "Metric": "Model Rows Loaded",
            "Value": len(combined_df),
        },
        {
            "Metric": "Missing / Skipped Inputs",
            "Value": len(missing_df),
        },
    ]

    if not combined_df.empty:
        ranked = combined_df.copy()

        ranked = ranked.sort_values(
            by=[
                "AverageBestRegularMatch_PerDraw",
                "DrawsWithAtLeast3RegularMatches",
                "AverageTotalScore_AllRows",
            ],
            ascending=[
                False,
                False,
                False,
            ]
        ).reset_index(drop=True)

        top = ranked.iloc[0]

        rows.extend([
            {
                "Metric": "Best Overall Game",
                "Value": top["GameFamily"],
            },
            {
                "Metric": "Best Overall Model",
                "Value": top["ModelName"],
            },
            {
                "Metric": "Best Avg Regular Match / Draw",
                "Value": round(
                    float(
                        top["AverageBestRegularMatch_PerDraw"]
                    ),
                    4
                ),
            },
        ])

    return pd.DataFrame(rows)


def build_unified_leaderboard(combined_df):
    if combined_df.empty:
        return pd.DataFrame()

    leaderboard = combined_df.copy()

    numeric_cols = [
        "AverageBestRegularMatch_PerDraw",
        "DrawsWithAtLeast3RegularMatches",
        "AverageTotalScore_AllRows",
        "BonusHitDrawRate",
    ]

    for col in numeric_cols:
        leaderboard = safe_numeric(
            leaderboard,
            col
        )

    sort_cols = [
        col for col in [
            "AverageBestRegularMatch_PerDraw",
            "DrawsWithAtLeast3RegularMatches",
            "AverageTotalScore_AllRows",
            "BonusHitDrawRate",
        ]
        if col in leaderboard.columns
    ]

    leaderboard = leaderboard.sort_values(
        by=sort_cols,
        ascending=[False] * len(sort_cols)
    ).reset_index(drop=True)

    leaderboard["UnifiedRank"] = leaderboard.index + 1

    preferred_cols = [
        "UnifiedRank",
        "GameFamily",
        "Rank",
        "ModelName",
        "PredictionRows",
        "DrawsTested",
        "AverageRegularMatches_AllRows",
        "AverageTotalScore_AllRows",
        "BestRegularMatch_AnyRow",
        "BestTotalScore_AnyRow",
        "AverageBestScore_PerDraw",
        "AverageBestRegularMatch_PerDraw",
        "DrawsWithAtLeast2RegularMatches",
        "DrawsWithAtLeast3RegularMatches",
        "DrawsWithBonusHit",
        "BonusHitDrawRate",
    ]

    cols = [
        col for col in preferred_cols
        if col in leaderboard.columns
    ]

    return leaderboard[cols]


def build_best_by_game(combined_df):
    if combined_df.empty:
        return pd.DataFrame()

    leaderboard = build_unified_leaderboard(
        combined_df
    )

    if leaderboard.empty:
        return leaderboard

    best = (
        leaderboard
        .sort_values(
            by=[
                "GameFamily",
                "AverageBestRegularMatch_PerDraw",
                "DrawsWithAtLeast3RegularMatches",
            ],
            ascending=[
                True,
                False,
                False,
            ]
        )
        .groupby("GameFamily")
        .head(1)
        .reset_index(drop=True)
    )

    return best


def build_vs_random(combined_df):
    if combined_df.empty:
        return pd.DataFrame()

    rows = []

    for game_family, group in combined_df.groupby(
        "GameFamily"
    ):
        random_rows = group[
            group["ModelName"] == "Random_Baseline"
        ]

        if random_rows.empty:
            continue

        random_row = random_rows.iloc[0]

        for _, row in group.iterrows():
            rows.append({
                "GameFamily": game_family,
                "ModelName": row["ModelName"],
                "AvgBestRegularMatch_PerDraw": row.get(
                    "AverageBestRegularMatch_PerDraw",
                    None
                ),
                "Random_AvgBestRegularMatch_PerDraw": random_row.get(
                    "AverageBestRegularMatch_PerDraw",
                    None
                ),
                "Difference_vs_Random": round(
                    float(row.get(
                        "AverageBestRegularMatch_PerDraw",
                        0
                    ))
                    - float(random_row.get(
                        "AverageBestRegularMatch_PerDraw",
                        0
                    )),
                    4
                ),
                "AtLeast3_Diff_vs_Random": int(
                    row.get(
                        "DrawsWithAtLeast3RegularMatches",
                        0
                    )
                    - random_row.get(
                        "DrawsWithAtLeast3RegularMatches",
                        0
                    )
                ),
                "BeatsRandom_AvgBestRegular": (
                    "Yes"
                    if float(row.get(
                        "AverageBestRegularMatch_PerDraw",
                        0
                    ))
                    > float(random_row.get(
                        "AverageBestRegularMatch_PerDraw",
                        0
                    ))
                    else "No"
                ),
            })

    result = pd.DataFrame(rows)

    if result.empty:
        return result

    result = result.sort_values(
        by=[
            "GameFamily",
            "Difference_vs_Random",
        ],
        ascending=[
            True,
            False,
        ]
    ).reset_index(drop=True)

    return result


def build_game_summary(combined_df):
    if combined_df.empty:
        return pd.DataFrame()

    summary = (
        combined_df
        .groupby("GameFamily")
        .agg(
            ModelsCompared=("ModelName", "nunique"),
            AvgRegularMatchAcrossModels=(
                "AverageBestRegularMatch_PerDraw",
                "mean"
            ),
            BestRegularMatchAcrossModels=(
                "AverageBestRegularMatch_PerDraw",
                "max"
            ),
            TotalDrawsTested=("DrawsTested", "max"),
        )
        .reset_index()
    )

    return summary


def build_notes():
    notes = [
        {
            "Section": "Purpose",
            "Note": "This dashboard combines model comparison outputs across all lottery games.",
        },
        {
            "Section": "Primary Metric",
            "Note": "AverageBestRegularMatch_PerDraw is the main comparison metric.",
        },
        {
            "Section": "Random Baseline",
            "Note": "Every model must be compared against Random_Baseline.",
        },
        {
            "Section": "UK49s Warning",
            "Note": "UK49s currently has limited history, so results are experimental until more historical data is loaded.",
        },
        {
            "Section": "Important",
            "Note": "Lottery systems are random. These outputs measure historical model behaviour only.",
        },
    ]

    return pd.DataFrame(notes)


# =========================================================
# EXPORT
# =========================================================

def export_unified_model_performance_dashboard():
    combined_df, missing_df = load_all_summaries()

    dashboard_summary = build_dashboard_summary(
        combined_df,
        missing_df
    )

    leaderboard = build_unified_leaderboard(
        combined_df
    )

    best_by_game = build_best_by_game(
        combined_df
    )

    vs_random = build_vs_random(
        combined_df
    )

    game_summary = build_game_summary(
        combined_df
    )

    notes = build_notes()

    BACKTEST_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
        mode="w"
    ) as writer:

        dashboard_summary.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

        leaderboard.to_excel(
            writer,
            sheet_name="Unified_Leaderboard",
            index=False
        )

        best_by_game.to_excel(
            writer,
            sheet_name="Best_By_Game",
            index=False
        )

        vs_random.to_excel(
            writer,
            sheet_name="Vs_Random",
            index=False
        )

        game_summary.to_excel(
            writer,
            sheet_name="Game_Summary",
            index=False
        )

        if not missing_df.empty:
            missing_df.to_excel(
                writer,
                sheet_name="Missing_Inputs",
                index=False
            )

        notes.to_excel(
            writer,
            sheet_name="Notes",
            index=False
        )

        if not combined_df.empty:
            combined_df.to_excel(
                writer,
                sheet_name="Raw_Combined",
                index=False
            )

    style_workbook()

    print("\nUnified model performance dashboard exported.")
    print(f"File: {OUTPUT_FILE}")

    if not leaderboard.empty:
        print("\nTop unified models:")
        print(
            leaderboard[
                [
                    "UnifiedRank",
                    "GameFamily",
                    "ModelName",
                    "AverageBestRegularMatch_PerDraw",
                    "DrawsWithAtLeast3RegularMatches",
                ]
            ]
            .head(10)
            .to_string(index=False)
        )

    return {
        "dashboard_summary": dashboard_summary,
        "leaderboard": leaderboard,
        "best_by_game": best_by_game,
        "vs_random": vs_random,
        "game_summary": game_summary,
        "file": OUTPUT_FILE,
    }


# =========================================================
# STYLING
# =========================================================

def style_header(ws):
    fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


def auto_fit_columns(ws):
    for col_idx in range(
        1,
        ws.max_column + 1
    ):
        col_letter = get_column_letter(
            col_idx
        )

        max_length = 0

        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            col_letter
        ].width = min(
            max_length + 3,
            45
        )


def style_body(ws):
    thin_border = Border(
        bottom=Side(
            style="thin",
            color="D1D5DB"
        )
    )

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center"
            )


def add_conditional_formatting(ws):
    headers = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value
    }

    heat_cols = [
        "AverageBestRegularMatch_PerDraw",
        "AverageTotalScore_AllRows",
        "BonusHitDrawRate",
        "Difference_vs_Random",
        "AvgRegularMatchAcrossModels",
        "BestRegularMatchAcrossModels",
    ]

    for col_name in heat_cols:
        if col_name not in headers:
            continue

        col_idx = headers[col_name]
        col_letter = get_column_letter(
            col_idx
        )

        if ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="FCA5A5",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FEF3C7",
                    end_type="max",
                    end_color="86EFAC",
                )
            )

    bar_cols = [
        "DrawsWithAtLeast2RegularMatches",
        "DrawsWithAtLeast3RegularMatches",
        "DrawsWithBonusHit",
        "ModelsCompared",
    ]

    for col_name in bar_cols:
        if col_name not in headers:
            continue

        col_idx = headers[col_name]
        col_letter = get_column_letter(
            col_idx
        )

        if ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="60A5FA"
                )
            )


def style_workbook():
    wb = load_workbook(
        OUTPUT_FILE
    )

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        style_header(ws)
        style_body(ws)
        auto_fit_columns(ws)
        add_conditional_formatting(ws)

    wb.save(
        OUTPUT_FILE
    )


# =========================================================
# CLI
# =========================================================

def main():
    export_unified_model_performance_dashboard()


if __name__ == "__main__":
    main()