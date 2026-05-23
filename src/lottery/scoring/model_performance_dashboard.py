from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


# =========================================================
# PROJECT PATHS
# =========================================================

BASE_DIR = Path(__file__).resolve().parents[3]

BACKTEST_DIR = BASE_DIR / "data" / "exports" / "backtesting"

POWERBALL_COMPARISON_FILE = BACKTEST_DIR / "powerball_model_comparison_backtest.xlsx"
POWERBALL_BASIC_BACKTEST_FILE = BACKTEST_DIR / "powerball_backtest_results.xlsx"

DASHBOARD_FILE = BACKTEST_DIR / "model_performance_dashboard.xlsx"


# =========================================================
# LOADERS
# =========================================================

def read_excel_sheet(path, sheet_name):
    if not path.exists():
        return pd.DataFrame()

    try:
        return pd.read_excel(
            path,
            sheet_name=sheet_name,
            engine="openpyxl"
        )
    except Exception:
        return pd.DataFrame()


def load_powerball_comparison():
    summary = read_excel_sheet(
        POWERBALL_COMPARISON_FILE,
        "Summary"
    )

    rank_summary = read_excel_sheet(
        POWERBALL_COMPARISON_FILE,
        "Rank_Summary"
    )

    hit_distribution = read_excel_sheet(
        POWERBALL_COMPARISON_FILE,
        "Hit_Distribution"
    )

    model_configs = read_excel_sheet(
        POWERBALL_COMPARISON_FILE,
        "Model_Configs"
    )

    detailed_results = read_excel_sheet(
        POWERBALL_COMPARISON_FILE,
        "Detailed_Results"
    )

    return {
        "summary": summary,
        "rank_summary": rank_summary,
        "hit_distribution": hit_distribution,
        "model_configs": model_configs,
        "detailed_results": detailed_results,
    }


# =========================================================
# DASHBOARD TABLES
# =========================================================

def build_dashboard_summary(summary_df):
    if summary_df.empty:
        return pd.DataFrame([
            {
                "Metric": "Status",
                "Value": "No model summary data found.",
            }
        ])

    ranked = summary_df.copy()

    ranked = ranked.sort_values(
        by=[
            "AverageBestRegularMatch_PerDraw",
            "DrawsWithAtLeast3RegularMatches",
            "AverageTotalScore_AllRows",
        ],
        ascending=[False, False, False]
    ).reset_index(drop=True)

    best_model = ranked.iloc[0]

    random_row = ranked[
        ranked["ModelName"] == "Random_Baseline"
    ]

    if not random_row.empty:
        random_row = random_row.iloc[0]
        improvement_vs_random = (
            best_model["AverageBestRegularMatch_PerDraw"]
            - random_row["AverageBestRegularMatch_PerDraw"]
        )
    else:
        improvement_vs_random = None

    rows = [
        {
            "Metric": "Report Generated At",
            "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        {
            "Metric": "Models Compared",
            "Value": summary_df["ModelName"].nunique(),
        },
        {
            "Metric": "Best Model",
            "Value": best_model["ModelName"],
        },
        {
            "Metric": "Best Model Avg Best Regular Match / Draw",
            "Value": round(best_model["AverageBestRegularMatch_PerDraw"], 4),
        },
        {
            "Metric": "Best Model 3+ Match Draws",
            "Value": int(best_model["DrawsWithAtLeast3RegularMatches"]),
        },
        {
            "Metric": "Best Model Bonus Hit Draw Rate",
            "Value": round(best_model["BonusHitDrawRate"], 4),
        },
    ]

    if improvement_vs_random is not None:
        rows.append({
            "Metric": "Best Model Improvement vs Random",
            "Value": round(improvement_vs_random, 4),
        })

    return pd.DataFrame(rows)


def build_leaderboard(summary_df):
    if summary_df.empty:
        return pd.DataFrame()

    leaderboard = summary_df.copy()

    leaderboard = leaderboard.sort_values(
        by=[
            "AverageBestRegularMatch_PerDraw",
            "DrawsWithAtLeast3RegularMatches",
            "AverageTotalScore_AllRows",
            "BonusHitDrawRate",
        ],
        ascending=[False, False, False, False]
    ).reset_index(drop=True)

    leaderboard["DashboardRank"] = leaderboard.index + 1

    preferred_cols = [
        "DashboardRank",
        "ModelName",
        "PredictionRows",
        "DrawsTested",
        "AverageRegularMatches_AllRows",
        "AverageTotalScore_AllRows",
        "BestRegularMatch_AnyRow",
        "BestTotalScore_AnyRow",
        "AverageBestScore_PerDraw",
        "AverageBestRegularMatch_PerDraw",
        "DrawsWithAtLeast2RegularMatches",
        "DrawsWithAtLeast3RegularMatches",
        "DrawsWithBonusHit",
        "BonusHitDrawRate",
    ]

    cols = [
        col for col in preferred_cols
        if col in leaderboard.columns
    ]

    return leaderboard[cols]


def build_random_comparison(summary_df):
    if summary_df.empty:
        return pd.DataFrame()

    random = summary_df[
        summary_df["ModelName"] == "Random_Baseline"
    ]

    if random.empty:
        return pd.DataFrame()

    random = random.iloc[0]

    rows = []

    for _, row in summary_df.iterrows():
        rows.append({
            "ModelName": row["ModelName"],
            "AvgBestRegularMatch_PerDraw": row["AverageBestRegularMatch_PerDraw"],
            "Random_AvgBestRegularMatch_PerDraw": random["AverageBestRegularMatch_PerDraw"],
            "Difference_vs_Random": round(
                row["AverageBestRegularMatch_PerDraw"]
                - random["AverageBestRegularMatch_PerDraw"],
                4
            ),
            "AtLeast2_Diff_vs_Random": int(
                row["DrawsWithAtLeast2RegularMatches"]
                - random["DrawsWithAtLeast2RegularMatches"]
            ),
            "AtLeast3_Diff_vs_Random": int(
                row["DrawsWithAtLeast3RegularMatches"]
                - random["DrawsWithAtLeast3RegularMatches"]
            ),
            "BonusHitRate_Diff_vs_Random": round(
                row["BonusHitDrawRate"]
                - random["BonusHitDrawRate"],
                4
            ),
            "BeatsRandom_AvgBestRegular": (
                "Yes"
                if row["AverageBestRegularMatch_PerDraw"]
                > random["AverageBestRegularMatch_PerDraw"]
                else "No"
            ),
        })

    comparison = pd.DataFrame(rows)

    comparison = comparison.sort_values(
        by="Difference_vs_Random",
        ascending=False
    ).reset_index(drop=True)

    return comparison


def build_hit_pivot(hit_distribution_df):
    if hit_distribution_df.empty:
        return pd.DataFrame()

    pivot = hit_distribution_df.pivot_table(
        index="ModelName",
        columns="RegularMatches",
        values="Count",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    pivot.columns = [
        f"Matches_{col}" if isinstance(col, int) else str(col)
        for col in pivot.columns
    ]

    return pivot


def build_rank_effectiveness(rank_summary_df):
    if rank_summary_df.empty:
        return pd.DataFrame()

    rank_df = rank_summary_df.copy()

    rank_df = rank_df.sort_values(
        by=["ModelName", "PredictionRank"]
    ).reset_index(drop=True)

    return rank_df


def build_model_notes():
    notes = [
        {
            "Section": "Purpose",
            "Note": "This dashboard compares PowerBall model performance against random baseline results.",
        },
        {
            "Section": "Main Question",
            "Note": "Does a statistical model outperform random number selection?",
        },
        {
            "Section": "Key Metric",
            "Note": "AverageBestRegularMatch_PerDraw is the primary practical metric.",
        },
        {
            "Section": "Secondary Metric",
            "Note": "DrawsWithAtLeast3RegularMatches shows stronger hit events.",
        },
        {
            "Section": "Warning",
            "Note": "Lottery outcomes are random. These models measure historical pattern behaviour, not guaranteed future wins.",
        },
        {
            "Section": "Next Improvement",
            "Note": "Use leaderboard results to tune model configuration, then rerun backtesting.",
        },
    ]

    return pd.DataFrame(notes)


# =========================================================
# EXPORT
# =========================================================

def export_model_performance_dashboard():
    data = load_powerball_comparison()

    summary_df = data["summary"]
    rank_summary_df = data["rank_summary"]
    hit_distribution_df = data["hit_distribution"]
    model_configs_df = data["model_configs"]
    detailed_results_df = data["detailed_results"]

    dashboard_summary = build_dashboard_summary(summary_df)
    leaderboard = build_leaderboard(summary_df)
    random_comparison = build_random_comparison(summary_df)
    hit_pivot = build_hit_pivot(hit_distribution_df)
    rank_effectiveness = build_rank_effectiveness(rank_summary_df)
    model_notes = build_model_notes()

    DASHBOARD_FILE.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    with pd.ExcelWriter(
        DASHBOARD_FILE,
        engine="openpyxl",
        mode="w"
    ) as writer:
        dashboard_summary.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

        leaderboard.to_excel(
            writer,
            sheet_name="Leaderboard",
            index=False
        )

        random_comparison.to_excel(
            writer,
            sheet_name="Vs_Random",
            index=False
        )

        hit_pivot.to_excel(
            writer,
            sheet_name="Hit_Distribution",
            index=False
        )

        rank_effectiveness.to_excel(
            writer,
            sheet_name="Rank_Effectiveness",
            index=False
        )

        model_configs_df.to_excel(
            writer,
            sheet_name="Model_Configs",
            index=False
        )

        model_notes.to_excel(
            writer,
            sheet_name="Notes",
            index=False
        )

        if not detailed_results_df.empty:
            detailed_results_df.head(5000).to_excel(
                writer,
                sheet_name="Detailed_Sample",
                index=False
            )

    style_dashboard_workbook()

    print("\nModel performance dashboard exported.")
    print(f"File: {DASHBOARD_FILE}")

    if not leaderboard.empty:
        print("\nTop models:")
        print(
            leaderboard[
                [
                    "DashboardRank",
                    "ModelName",
                    "AverageBestRegularMatch_PerDraw",
                    "DrawsWithAtLeast3RegularMatches",
                    "BonusHitDrawRate",
                ]
            ].head(10).to_string(index=False)
        )

    return {
        "dashboard_summary": dashboard_summary,
        "leaderboard": leaderboard,
        "random_comparison": random_comparison,
        "file": DASHBOARD_FILE,
    }


# =========================================================
# STYLING
# =========================================================

def style_header(ws):
    fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


def auto_fit_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)

        max_length = 0

        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[col_letter].width = min(
            max_length + 3,
            45
        )


def style_body(ws):
    thin_border = Border(
        bottom=Side(
            style="thin",
            color="D1D5DB"
        )
    )

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center"
            )


def add_conditional_formatting(ws):
    headers = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value
    }

    possible_heat_cols = [
        "AverageBestRegularMatch_PerDraw",
        "AverageTotalScore_AllRows",
        "BonusHitDrawRate",
        "Difference_vs_Random",
        "AvgRegularMatches",
        "AvgTotalScore",
    ]

    for col_name in possible_heat_cols:
        if col_name not in headers:
            continue

        col_idx = headers[col_name]
        col_letter = get_column_letter(col_idx)

        if ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="FCA5A5",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FEF3C7",
                    end_type="max",
                    end_color="86EFAC",
                )
            )

    bar_cols = [
        "DrawsWithAtLeast2RegularMatches",
        "DrawsWithAtLeast3RegularMatches",
        "DrawsWithBonusHit",
    ]

    for col_name in bar_cols:
        if col_name not in headers:
            continue

        col_idx = headers[col_name]
        col_letter = get_column_letter(col_idx)

        if ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="60A5FA"
                )
            )


def style_dashboard_workbook():
    wb = load_workbook(DASHBOARD_FILE)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        style_header(ws)
        style_body(ws)
        auto_fit_columns(ws)
        add_conditional_formatting(ws)

    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 45

        for cell in ws["A"]:
            cell.font = Font(
                bold=True,
                color="111827"
            )

    wb.save(DASHBOARD_FILE)


# =========================================================
# CLI
# =========================================================

def main():
    export_model_performance_dashboard()


if __name__ == "__main__":
    main()