from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from src.lottery.analytics.historical_schema import (
    MASTER_FILE,
    MASTER_DIR,
    HISTORICAL_COLUMNS,
    UPDATE_LOG_COLUMNS,
    standardise_historical_df,
    build_summary,
    build_update_log_row,
    empty_historical_df,
    empty_update_log_df,
)


# =========================================================
# SHEET NAMES
# =========================================================

HISTORICAL_SHEET = "Historical_Results"
SUMMARY_SHEET = "Summary"
UPDATE_LOG_SHEET = "Update_Log"
DATA_DICTIONARY_SHEET = "Data_Dictionary"


# =========================================================
# CREATE MASTER FILE IF MISSING
# =========================================================

def ensure_master_file_exists():
    MASTER_DIR.mkdir(parents=True, exist_ok=True)

    if MASTER_FILE.exists():
        return

    wb = Workbook()

    ws = wb.active
    ws.title = HISTORICAL_SHEET

    wb.create_sheet(SUMMARY_SHEET)
    wb.create_sheet(UPDATE_LOG_SHEET)
    wb.create_sheet(DATA_DICTIONARY_SHEET)

    wb.save(MASTER_FILE)


# =========================================================
# READ SHEETS
# =========================================================

def read_sheet(sheet_name, columns):
    ensure_master_file_exists()

    try:
        df = pd.read_excel(
            MASTER_FILE,
            sheet_name=sheet_name,
            engine="openpyxl"
        )

        if df.empty:
            return pd.DataFrame(columns=columns)

        return df

    except Exception:
        return pd.DataFrame(columns=columns)


def read_existing_history():
    return read_sheet(HISTORICAL_SHEET, HISTORICAL_COLUMNS)


def read_existing_update_log():
    return read_sheet(UPDATE_LOG_SHEET, UPDATE_LOG_COLUMNS)


# =========================================================
# DATA DICTIONARY
# =========================================================

def build_data_dictionary():
    rows = [
        {
            "Column": "GameFamily",
            "Description": "High-level lottery group.",
            "Example": "PowerBall",
        },
        {
            "Column": "GameName",
            "Description": "Specific game name.",
            "Example": "PowerBall Plus",
        },
        {
            "Column": "DrawType",
            "Description": "Main, Plus, Plus 1, Plus 2, Lunchtime, Teatime, etc.",
            "Example": "Plus",
        },
        {
            "Column": "DrawNumber",
            "Description": "Official draw number if available from the source.",
            "Example": "1623",
        },
        {
            "Column": "DrawDate",
            "Description": "Date of the draw.",
            "Example": "2026-05-10",
        },
        {
            "Column": "DrawDay",
            "Description": "Weekday of the draw.",
            "Example": "Friday",
        },
        {
            "Column": "N1 to N6",
            "Description": "Regular winning numbers. N6 is optional depending on game.",
            "Example": "5, 11, 23, 29, 48",
        },
        {
            "Column": "Bonus",
            "Description": "Bonus number / PowerBall number where applicable.",
            "Example": "8",
        },
        {
            "Column": "Jackpot",
            "Description": "Advertised jackpot value where available.",
            "Example": "72000000",
        },
        {
            "Column": "Outcome",
            "Description": "Outcome text from the source where available.",
            "Example": "Roll",
        },
        {
            "Column": "SourceName",
            "Description": "Website or provider used.",
            "Example": "za.national-lottery.com",
        },
        {
            "Column": "SourceUrl",
            "Description": "Exact URL used for the row.",
            "Example": "https://za.national-lottery.com/powerball/results/2026-archive",
        },
        {
            "Column": "LoadedAt",
            "Description": "Timestamp when the row was loaded into the master file.",
            "Example": "2026-05-10 12:00:00",
        },
        {
            "Column": "RecordKey",
            "Description": "Unique key used to prevent duplicate result rows.",
            "Example": "PowerBall|PowerBall|Main|2026-05-10|...",
        },
    ]

    return pd.DataFrame(rows)


# =========================================================
# WRITE MASTER WORKBOOK
# =========================================================

def write_master_workbook(history_df, summary_df, update_log_df, dictionary_df):
    MASTER_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(MASTER_FILE, engine="openpyxl", mode="w") as writer:
        history_df.to_excel(writer, sheet_name=HISTORICAL_SHEET, index=False)
        summary_df.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
        update_log_df.to_excel(writer, sheet_name=UPDATE_LOG_SHEET, index=False)
        dictionary_df.to_excel(writer, sheet_name=DATA_DICTIONARY_SHEET, index=False)

    style_master_workbook()


# =========================================================
# APPEND NEW HISTORY
# =========================================================

def append_history_rows(
    new_rows,
    update_type,
    game_family,
    game_name,
    draw_type,
    notes="",
):
    existing_history = read_existing_history()
    existing_log = read_existing_update_log()

    new_df = pd.DataFrame(new_rows)

    new_df = standardise_historical_df(new_df)
    existing_history = standardise_historical_df(existing_history)

    rows_fetched = len(new_df)

    if new_df.empty:
        update_row = build_update_log_row(
            update_type=update_type,
            game_family=game_family,
            game_name=game_name,
            draw_type=draw_type,
            rows_fetched=0,
            rows_added=0,
            rows_skipped=0,
            status="No Data",
            notes="No valid rows received. " + notes,
        )

        update_log = pd.concat(
            [existing_log, pd.DataFrame([update_row])],
            ignore_index=True
        )

        summary = build_summary(existing_history)
        dictionary = build_data_dictionary()

        write_master_workbook(existing_history, summary, update_log, dictionary)

        print("No valid rows received.")
        return {
            "rows_fetched": 0,
            "rows_added": 0,
            "rows_skipped": 0,
            "status": "No Data",
        }

    combined = pd.concat(
        [existing_history, new_df],
        ignore_index=True
    )

    before_dedup = len(combined)

    combined = standardise_historical_df(combined)

    after_dedup = len(combined)

    rows_added = max(after_dedup - len(existing_history), 0)
    rows_skipped = before_dedup - after_dedup

    update_row = build_update_log_row(
        update_type=update_type,
        game_family=game_family,
        game_name=game_name,
        draw_type=draw_type,
        rows_fetched=rows_fetched,
        rows_added=rows_added,
        rows_skipped=rows_skipped,
        status="Success",
        notes=notes,
    )

    update_log = pd.concat(
        [existing_log, pd.DataFrame([update_row])],
        ignore_index=True
    )

    summary = build_summary(combined)
    dictionary = build_data_dictionary()

    write_master_workbook(combined, summary, update_log, dictionary)

    print("Historical master updated.")
    print(f"File: {MASTER_FILE}")
    print(f"Rows fetched: {rows_fetched}")
    print(f"Rows added: {rows_added}")
    print(f"Rows skipped: {rows_skipped}")

    return {
        "rows_fetched": rows_fetched,
        "rows_added": rows_added,
        "rows_skipped": rows_skipped,
        "status": "Success",
    }


# =========================================================
# EXCEL STYLING
# =========================================================

def remove_existing_tables(ws):
    if ws.tables:
        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]


def style_header_row(ws, row_number=1):
    header_fill = PatternFill("solid", fgColor="FF2F7D")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_body(ws):
    thin_border = Border(
        bottom=Side(style="thin", color="334155")
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit_columns(ws, max_width=45):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 3, max_width)


def add_excel_table(ws, table_name):
    remove_existing_tables(ws)

    if ws.max_row < 2 or ws.max_column < 1:
        return

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    table = Table(displayName=table_name, ref=table_ref)

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)


def style_master_workbook():
    wb = load_workbook(MASTER_FILE)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        style_header_row(ws, 1)
        style_body(ws)
        autofit_columns(ws)

    if HISTORICAL_SHEET in wb.sheetnames:
        ws = wb[HISTORICAL_SHEET]
        add_excel_table(ws, "tblHistoricalResults")

        widths = {
            "A": 16,
            "B": 20,
            "C": 16,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 8,
            "H": 8,
            "I": 8,
            "J": 8,
            "K": 8,
            "L": 8,
            "M": 10,
            "N": 16,
            "O": 18,
            "P": 28,
            "Q": 65,
            "R": 22,
            "S": 70,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    if SUMMARY_SHEET in wb.sheetnames:
        ws = wb[SUMMARY_SHEET]
        add_excel_table(ws, "tblSummary")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

    if UPDATE_LOG_SHEET in wb.sheetnames:
        ws = wb[UPDATE_LOG_SHEET]
        add_excel_table(ws, "tblUpdateLog")

        widths = {
            "A": 24,
            "B": 22,
            "C": 18,
            "D": 22,
            "E": 18,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 16,
            "J": 70,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    if DATA_DICTIONARY_SHEET in wb.sheetnames:
        ws = wb[DATA_DICTIONARY_SHEET]
        add_excel_table(ws, "tblDataDictionary")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 75
        ws.column_dimensions["C"].width = 45

    wb.save(MASTER_FILE)


# =========================================================
# QUICK TEST
# =========================================================

if __name__ == "__main__":
    test_rows = [
        {
            "GameFamily": "PowerBall",
            "GameName": "PowerBall",
            "DrawType": "Main",
            "DrawNumber": "",
            "DrawDate": "2026-05-10",
            "DrawDay": "Sunday",
            "N1": 5,
            "N2": 11,
            "N3": 23,
            "N4": 29,
            "N5": 48,
            "N6": None,
            "Bonus": 1,
            "Jackpot": 25000000,
            "Outcome": "Roll",
            "SourceName": "Test",
            "SourceUrl": "https://example.com",
            "LoadedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "RecordKey": "",
        }
    ]

    append_history_rows(
        new_rows=test_rows,
        update_type="Test Insert",
        game_family="PowerBall",
        game_name="PowerBall",
        draw_type="Main",
        notes="Testing lottery master writer."
    )